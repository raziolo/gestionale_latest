import json
from datetime import datetime, timedelta
from pprint import pprint

from django.utils.datastructures import MultiValueDictKeyError

from django.shortcuts import render, reverse, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpRequest
from openpyxl import load_workbook
from io import BytesIO

from api.models import Import, Schedule

from api.models import Branch

from api import formulas as f


# Create your views here.
# views.py
from datetime import datetime, timedelta
from django.shortcuts import render
from api.models import Branch
# Import your formulas functions (adjust the import path as needed)

from datetime import datetime, timedelta
import logging

from django.shortcuts import render
from django.http import HttpResponse

# Set up logging (this example uses the standard library's logging)
logger = logging.getLogger(__name__)


def dashboard(request):
    date_param = None
    date = None
    date_start, date_end = None, None

    branches = Branch.objects.all()

    branches_om = []
    branches_equivalenza = []

    for branch in branches:
        if branch.get_brand() == "original":
            branches_om.append(branch)
        elif branch.get_brand() == "equivalenza":
            branches_equivalenza.append(branch)

    if request.method == "POST":
        date_param = request.POST.get('date', None)
        date = date_param

    # Get the query string parameters
    if not date_param:
        date_start_param = datetime.now().date() - timedelta(days=385)
        date_end_param = datetime.now().date() - timedelta(days=355)

        date_start_str = date_start_param.strftime('%Y-%m-%d')
        date_end_str = date_end_param.strftime('%Y-%m-%d')

        date = "{} to {}".format(date_start_str, date_end_str)

    # Convert the date parameter to string (it may be None)

    # Initialize date_start and date_end

    # Check if the date string contains a range (using "to")
    if "to" in date:
        parts = date.split("to")
        if len(parts) >= 2:
            date_start = parts[0].strip()
            date_end = parts[1].strip()
        else:
            date_start = date.strip()
            date_end = date.strip()
    else:
        # If no range is provided, you could either treat it as a single date or return an error.
        # Here, we treat it as a single date.
        date_start = date.strip()
        date_end = date.strip()

    start_date = datetime.strptime(date_start, '%Y-%m-%d')
    end_date = datetime.strptime(date_end, '%Y-%m-%d')

    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    branches_data = []
    for branch in branches:
        branch_data = {}
        # --- Compute All KPI Metrics using your formulas ---
        try:
            sales_data = f.generate_branch_report_sales(branch.id, start_date_str, end_date_str)
            receipts_data = f.generate_branch_report_scontrini(branch.id, start_date_str, end_date_str)
            traffic_data = f.generate_ingressi_branch_report(branch.id, start_date_str, end_date_str)
            external_traffic = f.generate_branch_traffico_esterno_report(branch.id, start_date_str, end_date_str)
            employee_performance = f.generate_report_performance_sales(branch.id, start_date_str, end_date_str)
            conversion_rate = f.generate_branch_report_conversion_rate(branch.id, start_date_str, end_date_str)

            categories = sorted(sales_data.keys(), reverse=False)

            branch_data.update({
                'id': branch.id,
                'name': branch.name,
                'brand': branch.get_brand(),
                'sales': sales_data,
                'receipts': receipts_data,
                'traffic': traffic_data,
                'conversion_rate': conversion_rate,
                'external_traffic': external_traffic,
                'employee_performance': employee_performance,
                'categories': categories,

            })
            branches_data.append(branch_data)

        except Exception as e:
            print(e)
            return HttpResponse("Error calculating metrics", status=500)



    context = {
        'branches_om': branches_om,
        'branches_equivalenza': branches_equivalenza,
        'branches_data': branches_data
    }

    pprint(branches_data)

    # If this is an HTMX request, return only the dashboard content partial
    if request.headers.get("HX-Request") == "true":
        print("htmx")
        return render(request, "frontend/dashboard_content.html", context)
    else:
        return render(request, "frontend/dashboard.html", context)


from django_tables2 import SingleTableView
from django_filters.views import FilterView

from .tables import EmployeeTable, EmployeePerformancesTable, SchedulesTable
from .filters import EmployeeFilter
from api.models import Employee

class EmployeeListView(FilterView, SingleTableView):
    model = Employee
    table_class = EmployeeTable
    template_name = "frontend/employees/all.html"
    filterset_class = EmployeeFilter

    def get_queryset(self):
        return Employee.objects.all()


def import_data(request):
    if request.method == "GET":
        return render(request, "frontend/import/import.html")
    if request.method == "POST":
        # Get the uploaded file
        uploaded_file = request.FILES.get('file')
        selected_branch = request.POST.get('branchSelect')
        selected_type = request.POST.get('typeSelect')



        try:
            selected_branch = int(selected_branch)
        except:
            return JsonResponse({"status": "error", "errors": ["Internal"]}, status=200)

        if not selected_branch:
            return JsonResponse({"status": "error", "errors": ["no branch selected"]}, status=200)

        data_dict = dict

        branch_obj = Branch.objects.get(id=selected_branch)

        if branch_obj.get_brand() == "equivalenza":

            if selected_type == "counter_data":
                if uploaded_file:
                    # Read the file as binary
                    file_content = uploaded_file.read()

                    # Load the workbook from the binary data
                    workbook = load_workbook(filename=BytesIO(file_content))
                    sheet = workbook.active  # You can specify sheet name if needed

                    # Initialize the dictionary to hold the data
                    data_dict = {}

                    errors = []

                    total_rows = sheet.max_row

                    # Iterate through the rows, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                        date = row[0]  # assuming the date is in the second column
                        record = {
                            "(Ing) Ingressi": row[1] or 0,
                            "(Est) Traffico Esterno": row[2] or 0,
                            "(TA) Tasso di Attrazione": row[3] or 0,
                        }

                        # mar-19.11.24 to YYYY/MM/DD
                        date = date.split("-")[1]
                        date = datetime.strptime(date, "%d.%m.%y").strftime("%Y-%m-%d")


                        # If the date is already a key, overwrite the current record
                        if date in data_dict:

                            data_dict[date] = record
                        else:
                            # Otherwise, create a new list with the record
                            data_dict[date] = [record]


                        data_dict[date] = [record]


                    for date, records in data_dict.items():
                        if Import.objects.filter(branch=branch_obj, import_date=date, import_type=selected_type).exists():
                            errors.append(f"Collision on {date}")
                            continue


                    if errors:
                        return JsonResponse({"status": "error", "errors": errors}, status=200)

                    import_bulk_create_list = []
                    for date, data in data_dict.items():
                        i = Import(import_date=date, data=data, branch=branch_obj, import_type=selected_type)
                        import_bulk_create_list.append(i)

                    Import.objects.bulk_create(import_bulk_create_list)

                    return JsonResponse({"status": "success"}, status=200)

            ### START CONVERTING DATA
            elif selected_type == "sales_data":
                if uploaded_file:
                    # Read the file as binary
                    file_content = uploaded_file.read()

                    # Load the workbook from the binary data
                    workbook = load_workbook(filename=BytesIO(file_content))
                    sheet = workbook.active  # You can specify sheet name if needed

                    # Initialize the dictionary to hold the data
                    data_dict = {}

                    # Iterate through the rows, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                        date = row[1]  # assuming the date is in the second column
                        record = {
                            "Dipendente": row[0],
                            "Qta. Vend.": row[2],
                            "Sco.": row[3],
                            "Importo": row[4],
                            "Sco. Medio": row[5],
                            "Qta Media": row[6]
                        }

                        # If the date is already a key, append the new record to its list
                        if date in data_dict:
                            data_dict[date].append(record)
                        else:
                            # Otherwise, create a new list with the record
                            data_dict[date] = [record]

                    import_qs = Import.objects.none()

                    errors = []

                    for date, records in data_dict.items():
                        employees_day_id_list = []
                        date = datetime.strptime(date, '%d/%m/%Y').strftime('%Y-%m-%d')
                        for record in records:
                            try:
                                employees_day_id_list.append(record['Dipendente'])
                            except:
                                errors.append(f"Employee {record['Dipendente']} not found")
                                continue

                        employees_day_qs = Employee.objects.filter(id__in=employees_day_id_list)

                        if employees_day_qs.count() != len(employees_day_id_list):
                            errors.append(f"Employees on {date} not found")
                            continue

                        branch_check_value = list(employees_day_qs.values_list('branch', flat=True))
                        normalized = list(dict.fromkeys(branch_check_value))
                        if len(normalized) != 1:
                            errors.append(f"Branch on {date} from different branch")
                            continue

                        if normalized[0] != selected_branch:
                            errors.append(f"Branch on {date} from different branch")
                            continue

                        if Import.objects.filter(branch=branch_obj, import_date=date):
                            errors.append(f"Collision on {date}")
                            continue

                    if errors:
                        return JsonResponse({"status": "error", "errors": errors}, status=200)

                    import_bulk_create_list = []
                    for date, data in data_dict.items():
                        date_obj = datetime.strptime(date, '%d/%m/%Y')
                        date_str = date_obj.strftime('%Y-%m-%d')
                        i = Import(import_date=date_str, data=data, branch=branch_obj, import_type=selected_type)
                        import_bulk_create_list.append(i)

                    Import.objects.bulk_create(import_bulk_create_list)


                return JsonResponse({"status" : "success", "errors": []}, status=200)

        elif branch_obj.get_brand() == "original":
            print("original")

            if selected_type == "counter_data":
                if uploaded_file:
                    # Read the file as binary
                    file_content = uploaded_file.read()

                    # Load the workbook from the binary data
                    workbook = load_workbook(filename=BytesIO(file_content))
                    sheet = workbook.active  # You can specify sheet name if needed

                    # Initialize the dictionary to hold the data
                    data_dict = {}

                    errors = []

                    total_rows = sheet.max_row



                    # Iterate through the rows, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row



                        date = row[0]  # assuming the date is in the second column
                        record = {
                            "(Ing) Ingressi": row[1],
                            "(Est) Traffico Esterno": row[2],
                            "(TA) Tasso di Attrazione": row[3],
                        }

                        data_dict[date] = [record]

                    print(data_dict)


            ### START CONVERTING DATA
            elif selected_type == "sales_data":

                if uploaded_file:
                    # Read the file as binary
                    file_content = uploaded_file.read()

                    # Load the workbook from the binary data
                    workbook = load_workbook(filename=BytesIO(file_content))
                    sheet = workbook.active  # You can specify sheet name if needed

                    # Initialize the dictionary to hold the data
                    data_dict = {}

                    errors = []

                    total_rows = sheet.max_row

                    # Iterate through the rows, skipping the header
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row

                        if row[1] in ["Totale generale", "Totali "]:
                            continue


                        date = row[1]  # assuming the date is in the second column
                        record = {
                            "Qta. Vend.": row[2],
                            "Sco.": row[3],
                            "Importo": row[4],
                            "Sco. Medio": row[5],
                            "Qta Media": row[6]
                        }
                        if Import.objects.filter(branch=branch_obj, import_date=date):
                            errors.append(f"Collision on {date}")
                            continue

                        data_dict[date] = [record]

                    import_qs = Import.objects.none()

                    if errors:
                        return JsonResponse({"status": "error", "errors": errors}, status=200)

                    data_dict_formatted = {}

                    data_dict = {k: v for k, v in data_dict.items() if k is not None}

                    for k,v in data_dict.items():

                        unformatted = k
                        date_strp = datetime.strptime(unformatted, '%d/%m/%Y')
                        date_str = date_strp.strftime('%Y-%m-%d')
                        data_dict_formatted[date_str] = v

                    import_bulk_create_list = []
                    for date, data in data_dict_formatted.items():
                        i = Import(import_date=date, data=data, branch=branch_obj, import_type=selected_type)
                        import_bulk_create_list.append(i)

                    Import.objects.bulk_create(import_bulk_create_list)



                return JsonResponse({"status": "success", "errors": []}, status=200)



def filter(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        filter_type = data.get('type')
        branch = data.get('branch')
        date = data.get('date')
        print(filter_type, branch, date)  # Should now print your values

        redirect_url = ""
        if filter_type == "branch":

            redirect_url = "/report_branch?branch={}&date={}".format(branch, date)

        elif filter_type == "employees":

            redirect_url = "/report_employees?branch={}&date={}".format(branch, date)

        elif filter_type == "counter":

            redirect_url = "/report_counter?branch={}&date={}".format(branch, date)

        return JsonResponse({"redirect_url": redirect_url})

def report_branch(request):
    if request.method == 'GET':
        # Get the query string parameters
        branch_param = request.GET.get('branch')
        date_param = request.GET.get('date')

        # Validate branch parameter
        if not branch_param:
            return JsonResponse({"status": "error", "errors": ["No branch selected"]}, status=400)
        try:
            branch_id = int(branch_param)
        except ValueError:
            return JsonResponse({"status": "error", "errors": ["Invalid branch ID"]}, status=400)

        # Convert the date parameter to string (it may be None)
        date = str(date_param) if date_param else ""

        # Initialize date_start and date_end
        date_start, date_end = None, None

        # Check if the date string contains a range (using "to")
        if "to" in date:
            parts = date.split("to")
            if len(parts) >= 2:
                date_start = parts[0].strip()
                date_end = parts[1].strip()
            else:
                date_start = date.strip()
                date_end = date.strip()
        else:
            # If no range is provided, you could either treat it as a single date or return an error.
            # Here, we treat it as a single date.
            date_start = date.strip()
            date_end = date.strip()

        # Generate the report (adjust parameters as needed)
        sc = f.generate_branch_report_scontrini(branch_id, date_start, date_end)
        sc_total = int(sum(sc.values())) ## TOTALE SCONTRINI


        branch = Branch.objects.get(id=branch_id)

        sales = f.generate_branch_report_sales(branch_id, date_start, date_end)
        sales_total = sum(sales.values())
        sales_total = round(sales_total, 2)

        zoom_enabled = "false"
        graph_type = "area"

        if len(sales) > 30:
            zoom_enabled = "true"
            graph_type = "area"

        brand = branch.get_brand()
        print(brand)

        context = {
            "sc": sc,
            "sc_total": sc_total,
            "sales_total": sales_total,
            "brand": brand,
            "date_start": date_start,
            "branch": branch,
            "date_end": date_end,
            "sales": sales,
            "zoom_enabled": zoom_enabled,
            "type" : graph_type,
        }

        return render(request, "frontend/report/branch.html", context)

    # Optionally handle non-GET requests
    return JsonResponse({"status": "error", "errors": ["Invalid request method"]}, status=405)



def report_employees(request):
    if request.method == 'GET':
        # Get the query string parameters
        branch_param = request.GET.get('branch')
        date_param = request.GET.get('date')



        # Validate branch parameter
        if not branch_param:
            return JsonResponse({"status": "error", "errors": ["No branch selected"]}, status=400)
        try:
            branch_id = int(branch_param)
        except ValueError:
            return JsonResponse({"status": "error", "errors": ["Invalid branch ID"]}, status=400)

        try:
            branch = Branch.objects.get(id=branch_id)
        except Branch.DoesNotExist:
            return JsonResponse({"status": "error", "errors": ["Branch not found"]}, status=400)

        if branch.get_brand() == "original":
            return JsonResponse({"status": "error", "errors": ["Original Marines No Employee Performances"]}, status=400)

        # Convert the date parameter to string (it may be None)
        date = str(date_param) if date_param else ""

        # Initialize date_start and date_end
        date_start, date_end = None, None

        # Check if the date string contains a range (using "to")
        if "to" in date:
            parts = date.split("to")
            if len(parts) >= 2:
                date_start = parts[0].strip()
                date_end = parts[1].strip()
            else:
                date_start = date.strip()
                date_end = date.strip()
        else:
            # If no range is provided, you could either treat it as a single date or return an error.
            # Here, we treat it as a single date.
            date_start = date.strip()
            date_end = date.strip()

        # Generate the report (adjust parameters as needed)



        sc_performance = f.generate_report_performance_scontrini(branch_id, date_start, date_end)

        sales_performance = f.generate_report_performance_sales(branch_id, date_start, date_end)


        medium_sc_sales = f.generate_medium_sc_sales(sc_performance)

        medium_sales_performance = f.generate_medium_sales_performance(sales_performance)

        medium_number_sales_performance = f.generate_medium_sales(sales_performance)

        performance_table_data = []

        for key, value in sc_performance.items():
            performance_table_data.append({
                    "employee": key,
                    "quantita": int(medium_number_sales_performance[key]),
                    "n_scontrini": medium_sc_sales[key],
                    "importo": medium_sales_performance[key],
                }
            )

        performances_table = EmployeePerformancesTable(performance_table_data, orderable=False)

        zoom_enabled = "true"
        graph_type = "area"

        if date_start == date_end:
            zoom_enabled = "false"
            graph_type = "bar"

        total_sales = sum([sum(sales) for sales in sales_performance.values()])
        total_sc = sum([sum(sc) for sc in sc_performance.values()])

        # Calculate the percentage for each employee within each KPI
        sales_percentage = {key: round((sum(sales) / total_sales) * 100, 2) if total_sales != 0 else 0
                            for key, sales in sales_performance.items()}
        sc_percentage = {key: round((sum(sc) / total_sc) * 100, 2) if total_sc != 0 else 0
                         for key, sc in sc_performance.items()}


        context = {
            "sc_performance": sc_performance,
            "branch": Branch.objects.get(id=branch_id),
            "date_start": date_start,
            "date_end": date_end,
            "sales_performance": sales_performance,
            "performances_table": performances_table,
            "sales_percentage": sales_percentage,
            "sc_percentage": sc_percentage,
            "zoom_enabled": zoom_enabled,
            "type" : graph_type,
        }

        return render(request, "frontend/report/employees.html", context)


def config(request):
    if request.method == 'GET':
        return render(request, "frontend/base_cms/config.html")


def import_history(request):
    from datetime import timedelta, datetime

    if request.method == 'GET':
        current_year = datetime.now().year
        years_list = range(current_year, current_year - 10, -1)

        return render(request, "frontend/import/history.html", {
            'years_list': years_list,
            'year': current_year,
        })

    elif request.method == 'POST':
        form_data = request.POST

        branch_id = form_data.get('branchSelect')
        year = form_data.get('year')

        try:
            branch = Branch.objects.get(id=branch_id)
        except Branch.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Branch not found'})

        try:
            m_year = int(year)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid year'})

        imports_objs_qs = Import.objects.filter(branch=branch)
        imports_objs_qs_filtered = Import.objects.none()

        # Filter for all import_obj containing the requested year
        for import_obj in imports_objs_qs:
            if import_obj.import_date.__contains__(year):
                imports_objs_qs_filtered |= Import.objects.filter(import_date=import_obj.import_date)

        # Build dictionary of all days in the selected year
        start_date = datetime(int(year), 1, 1)
        end_date = datetime(int(year), 12, 31)
        all_days_in_year = [
            (start_date + timedelta(days=i)).strftime('%Y-%m-%d')
            for i in range((end_date - start_date).days + 1)
        ]

        imports_report_mapped = {}

        # Mark days with import as 1
        for import_obj in imports_objs_qs_filtered:
            imports_report_mapped[import_obj.import_date] = 1
            if import_obj.import_date in all_days_in_year:
                all_days_in_year.remove(import_obj.import_date)

        # Remaining days get 0 for no import
        for day in all_days_in_year:
            imports_report_mapped[day] = 0

        # Now mark FUTURE days as -1
        today_str = datetime.now().strftime('%Y-%m-%d')
        today_date = datetime.strptime(today_str, '%Y-%m-%d').date()

        for day_str in imports_report_mapped.keys():
            day_date = datetime.strptime(day_str, '%Y-%m-%d').date()
            if day_date > today_date:
                imports_report_mapped[day_str] = -1

        # Sort dictionary by date
        imports_report_mapped = dict(sorted(imports_report_mapped.items()))

        # Rebuild years_list for re-render
        current_year = datetime.now().year
        years_list = range(current_year, current_year - 10, -1)

        context = {
            'imports': imports_report_mapped,
            'branch': branch,
            'years_list': years_list,
            'year': year,
        }

        return render(request, "frontend/import/history.html", context=context)


from django.shortcuts import render, redirect
from .forms import EmployeeForm, ScheduleForm


def new_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('all_employees')  # Change this to whatever view or URL you want
    else:
        form = EmployeeForm()

    return render(request, 'frontend/employees/new.html', {'form': form})


def report_counter(request):
    if request.method == 'GET':
        # Get the query string parameters
        branch_param = request.GET.get('branch')
        date_param = request.GET.get('date')

        # Validate branch parameter
        if not branch_param:
            return JsonResponse({"status": "error", "errors": ["No branch selected"]}, status=400)
        try:
            branch_id = int(branch_param)
        except ValueError:
            return JsonResponse({"status": "error", "errors": ["Invalid branch ID"]}, status=400)

        # Convert the date parameter to string (it may be None)
        date = str(date_param) if date_param else ""

        # Initialize date_start and date_end
        date_start, date_end = None, None

        # Check if the date string contains a range (using "to")
        if "to" in date:
            parts = date.split("to")
            if len(parts) >= 2:
                date_start = parts[0].strip()
                date_end = parts[1].strip()
            else:
                date_start = date.strip()
                date_end = date.strip()
        else:
            # If no range is provided, you could either treat it as a single date or return an error.
            # Here, we treat it as a single date.
            date_start = date.strip()
            date_end = date.strip()


        # Generate the report (adjust parameters as needed)
        ingressi = f.generate_ingressi_branch_report(branch_id, date_start, date_end)
        ingressi_total = int(sum(ingressi.values()))  ## TOTALE SCONTRINI

        branch = Branch.objects.get(id=branch_id)

        attrazione = f.generate_branch_tasso_attrazione_report(branch_id, date_start, date_end)
        attrazione_total = sum(attrazione.values())  / len(attrazione)
        attrazione_total = round(attrazione_total, 2)

        traffico_esterno = f.generate_branch_traffico_esterno_report(branch_id, date_start, date_end)
        traffico_esterno_total = sum(traffico_esterno.values())

        pprint(ingressi)

        zoom_enabled = "false"
        graph_type = "area"

        if len(attrazione) > 30:
            zoom_enabled = "true"
            graph_type = "area"

        context = {
            "branch": branch,
            "brand": branch.get_brand(),

            "date_start": date_start,
            "date_end": date_end,

            "ingressi": ingressi,
            "ingressi_total": ingressi_total,
            "attrazione": attrazione,
            "attrazione_total": attrazione_total,
            "traffico_esterno": traffico_esterno,
            "traffico_esterno_total": traffico_esterno_total,

            "zoom_enabled": zoom_enabled,
            "type": graph_type,
        }

        return render(request, "frontend/report/counter.html", context=context)

    # Optionally handle non-GET requests
    return JsonResponse({"status": "error", "errors": ["Invalid request method"]}, status=405)


from django_tables2 import SingleTableView
from django_filters.views import FilterView
from .tables import SchedulesTable
from .filters import ScheduleFilter  # Make sure ScheduleFilter is imported


class ScheduleListView(FilterView, SingleTableView):
    model = Schedule
    table_class = SchedulesTable
    template_name = "frontend/schedules/all.html"
    filterset_class = ScheduleFilter  # This should be your filterset class

    # You do not need to override `get_queryset` because `FilterView` handles it
    # If you still need custom logic, you can override `get_filterset` or `get_queryset` accordingly.

    # If overriding `get_queryset`, make sure it's using the filter:
    def get_queryset(self):
        # Default filtering is handled by the FilterView, but you can still customize here if needed.
        return Schedule.objects.all()  # Return all schedules if no filters applied


def new_schedule(request):
    if request.method == 'POST':
        form = ScheduleForm(request.POST)

        if form.is_valid():
            # Save the Orario instance
            new_sch = form.save()
            request.session['new_schedule_pk'] = new_sch.pk
            return redirect("config_schedule") #### redirect here with parameter
        else:
            print(form.errors)


    else:
        form = ScheduleForm()

    # If you want to render the employees separately in the template,
    # you can pass them in context.
    # Build a mapping of branch id -> list of employees (id, name, surname)
    employees_by_branch = {}
    for branch in Branch.objects.all():
        emps = Employee.objects.filter(branch=branch)
        employees_by_branch[branch.id] = list(emps.values("id", "first_name", "last_name"))

    context = {
        'form': form,
        # Pass the mapping as JSON so we can use it in JS.
        'employees_by_branch': json.dumps(employees_by_branch),
    }
    return render(request, "frontend/schedules/new.html", context=context)


def config_schedule(request):
    if request.method == "GET":
        new_schedule_pk = request.session.get('new_schedule_pk')
        schedule = Schedule.objects.get(pk=new_schedule_pk)
        employees = []
        for employee in schedule.employees:
            employees.append(Employee.objects.get(id=employee))

        # Serialize the shift data if it exists, otherwise use an empty list.
        prefilled_shifts = json.dumps(schedule.shift_data) if schedule.shift_data else '[]'

        free_days_map = {}
        if schedule.free_days:
            for entry in schedule.free_days:
                # Use the employee_id as string for easier lookup from HTML (data attributes are strings)
                free_days_map[str(entry.get("employee_id"))] = entry.get("free_days", [])
        # Convert to JSON so we can use it in JavaScript
        free_days_json = json.dumps(free_days_map)

        context = {
            "schedule": schedule,
            "employees": employees,
            "prefilled_shifts": prefilled_shifts,  # Pass the JSON string to the template.
            "free_days": free_days_json,  # JSON string mapping employee_id to free days

        }
        return render(request, "frontend/schedules/config.html", context=context)

    elif request.method == "POST":
        form_data = request.POST
        print(form_data)
        # Get and parse shifts data from request
        shifts_data = request.POST.get('shifts_data', '[]')  # Default to empty list if no shifts
        shifts = json.loads(shifts_data)

        new_schedule_pk = request.session.get('new_schedule_pk')
        schedule = Schedule.objects.get(pk=new_schedule_pk)

        # Convert minutes to HH:MM format
        if shifts != schedule.shift_data:
            for shift in shifts:
                shift["start"] = str(timedelta(minutes=shift["start"]))[:-3]  # Converts to HH:MM
                shift["end"] = str(timedelta(minutes=shift["end"]))[:-3]
            schedule.shift_data = shifts

        schedule.free_days = process_free_days(form_data)

        schedule.save()

        return redirect("confirm_schedule")


import json


def process_free_days(form_data):
    """
    Extract free days from the POST data.
    Returns a list of dictionaries, e.g.:
    [
        {"employee_id": 1, "free_days": ["2025-02-01", "2025-02-07", "2025-02-14"]},
        {"employee_id": 2, "free_days": ["2025-02-12", "2025-02-19", "2025-02-26"]},
        # ...
    ]
    """
    free_days_data = []

    # Loop through all keys in form_data.
    for key in form_data:
        if key.startswith("free_days_"):
            # Extract the employee ID from the key.
            employee_id_str = key.split("_")[-1]
            try:
                employee_id = int(employee_id_str)
            except ValueError:
                # Skip if employee_id isn't an integer.
                continue

            # Get the value (a list with one string element).
            free_days_value = form_data.get(key, "").strip()

            # Split the dates by comma and remove extra spaces.
            if free_days_value:
                days_list = [day.strip() for day in free_days_value.split(",") if day.strip()]
            else:
                days_list = []

            free_days_data.append({
                "employee_id": employee_id,
                "free_days": days_list,
            })

    return free_days_data





def confirm_schedule(request):
    if request.method == "GET":
        new_schedule_pk = request.session.get('new_schedule_pk')
        schedule = Schedule.objects.get(pk=new_schedule_pk)

        # Build a mapping for free days:
        free_days_map = {}
        if schedule.free_days:
            for entry in schedule.free_days:
                free_days_map[str(entry.get("employee_id"))] = entry.get("free_days", [])

        # Build a list of employee objects and annotate with free_days:
        employees = []
        for employee_id in schedule.employees:
            try:
                emp = Employee.objects.get(id=employee_id)
                # Annotate the employee object with its free days:
                emp.free_days = free_days_map.get(str(emp.id), [])
                employees.append(emp)
            except Employee.DoesNotExist:
                pass

        context = {
            "schedule": schedule,
            "employees": employees,
            # You can still pass free_days mapping if needed for other purposes:
            "free_days": free_days_map,
        }
        return render(request, "frontend/schedules/confirm.html", context=context)


def create_schedule(request):
    if request.method == "POST":
        new_schedule_pk = request.session.get('new_schedule_pk')
        schedule = Schedule.objects.get(pk=new_schedule_pk)
        from frontend.orario_creation import create_scheduleMP

        result = create_scheduleMP(schedule.id)

        if result == 0:
            return JsonResponse({"success": True, "message": "Operazione completata con successo"}, status=200)
        else:
            return JsonResponse({"success": False, "message": "-1"}, status=400)

def timeline_schedule(request, schedule_id):
    # Fetch the schedule object
    schedule = get_object_or_404(Schedule, pk=schedule_id)

    # If schedule_data is a JSON string, parse it:
    schedule_data = schedule.schedule_data
    if isinstance(schedule_data, str):
        schedule_data = json.loads(schedule_data)

    # We'll build a flattened list of (day, minute_mark) pairs, sorted by day and time
    days = sorted(schedule_data.keys())  # e.g. ["2025-02-01", "2025-02-02", ...]
    columns = []
    for day in days:
        # Each day has a dict of half-hour keys; we sort them numerically
        day_times = sorted(schedule_data[day].keys(), key=lambda x: int(x))
        for t in day_times:
            columns.append((day, t))  # e.g. ("2025-02-01", "540")

    # For each day, get the sorted time keys
    day_times_map = {}
    for day in days:
        time_keys = sorted(schedule_data[day].keys(), key=lambda x: int(x))
        day_times_map[day] = time_keys

    # Build a flat list of (day, time_key) to iterate in the second header row and table body
    columns = []
    for day in days:
        for t in day_times_map[day]:
            columns.append((day, t))

    # List of employees for your table
    employees = []

    for employee in schedule.employees:
        emp = Employee.objects.get(id=employee)
        emp_data = {
            "id": emp.id,
            "name": f"{emp.first_name} {emp.last_name}",
        }
        employees.append(emp_data)

    context = {
        "schedule": schedule,  # We might need schedule.id if we do editing
        "schedule_data": schedule_data,  # day->(time->list_of_employee_ids)
        "sorted_days": days,  # For the first header row
        "day_times_map": day_times_map,  # day-> sorted list of half-hours
        "columns": columns,  # flattened (day, time_key)
        "employees": employees,
    }
    return render(request, "frontend/schedules/timeline.html", context)


def set_schedule_for_processing(request, schedule_id):
    if request.method == 'GET':
        schedule = Schedule.objects.get(pk=schedule_id)

        if not schedule.processed:
            request.session['new_schedule_pk'] = schedule.pk
            return redirect("config_schedule")
        else:
            return redirect("all_schedules")

def set_schedule_for_modify(request, schedule_id):
    if request.method == 'GET':
        schedule = Schedule.objects.get(pk=schedule_id)
        if not schedule.processed:
            request.session['new_schedule_pk'] = schedule.pk
            return redirect("config_schedule")
        else:
            return redirect("all_schedules")


def delete_schedule(request, schedule_id):
    if request.method == 'GET':
        schedule = Schedule.objects.get(pk=schedule_id)

        if not schedule.processed:
            schedule.delete()
            return redirect("all_schedules")

        return redirect("all_schedules")


@require_POST
def toggle_assignment_bulk(request):
    data = json.loads(request.body)
    schedule_id = data["schedule_id"]
    toggles = data["toggles"]  # list of { day, time_str, employee_id, assign }

    # Example logic:
    schedule = get_object_or_404(Schedule, pk=schedule_id)
    schedule_data = json.loads(schedule.schedule_data)  # or however it's stored

    for t in toggles:
        day = t["day"]
        time_str = t["time_str"]
        employee_id = str(t["employee_id"])  # convert to string to match stored IDs

        # Access the assigned list
        assigned = schedule_data[day][time_str]

        if employee_id in assigned:
            # Remove the assignment
            assigned.remove(employee_id)
            assigned_state = False
        else:
            # Add the assignment
            assigned.append(employee_id)
            assigned_state = True

    # Save the updated schedule data
    schedule.schedule_data = json.dumps(schedule_data)
    schedule.save()

    return JsonResponse({"success": True})


def worked_hours(request):
    if request.method == "GET":
        return render(request, "frontend/employees/worked_hours.html")
    if request.method == "POST":
        data = json.loads(request.body)
        branch_param = data.get('branch')
        date_param = data.get('date')


        # Validate branch parameter
        if not branch_param:
            return JsonResponse({"status": "error", "errors": ["No branch selected"]}, status=400)
        try:
            branch_id = int(branch_param)
        except ValueError:
            return JsonResponse({"status": "error", "errors": ["Invalid branch ID"]}, status=400)

        # Convert the date parameter to string (it may be None)
        date = str(date_param) if date_param else ""

        # Initialize date_start and date_end
        date_start, date_end = None, None

        # Check if the date string contains a range (using "to")
        if "to" in date:
            parts = date.split("to")
            if len(parts) >= 2:
                date_start = parts[0].strip()
                date_end = parts[1].strip()
            else:
                date_start = date.strip()
                date_end = date.strip()
        else:
            # If no range is provided, you could either treat it as a single date or return an error.
            # Here, we treat it as a single date.
            date_start = date.strip()
            date_end = date.strip()

        date_start_dt = datetime.strptime(date_start, "%Y-%m-%d").date()
        date_end_dt = datetime.strptime(date_end, "%Y-%m-%d").date()
        date_range = [date_start_dt + timedelta(days=i) for i in range((date_end_dt - date_start_dt).days + 1)]

        employees_branch = Employee.objects.filter(branch=branch_id)

        worked_hours_data = {date.strftime("%Y-%m-%d"):[] for date in date_range}
        for current_date in date_range:
            for employee in employees_branch:
                w = {"employee": f"({employee.id}) {employee.first_name} {employee.last_name}",
                     "worked_hours": f.get_employee_worked_hours_single_day(employee.id, current_date.strftime("%Y-%m-%d"))}
                worked_hours_data[current_date.strftime("%Y-%m-%d")].append(w)


        all_employees = set()
        for day_data in worked_hours_data.values():
            for rec in day_data:
                all_employees.add(rec["employee"])

        all_employees = sorted(all_employees)  # optional

        # day -> {employee_id -> hours}
        pivot_data = {}

        for day, records in worked_hours_data.items():
            pivot_data[day] = {}
            for emp in all_employees:
                pivot_data[day][emp] = 0  # default 0

            for rec in records:
                pivot_data[day][rec["employee"]] = rec["worked_hours"]

        totals = {}
        for day, emp_dict in pivot_data.items():
            for emp_id, hours in emp_dict.items():
                totals[emp_id] = totals.get(emp_id, 0) + hours

        # 2) Add this dictionary back under a special key, e.g. "TOTAL".
        pivot_data["TOTALE"] = totals


        return JsonResponse({"status": "success", "worked_hours_data": pivot_data})